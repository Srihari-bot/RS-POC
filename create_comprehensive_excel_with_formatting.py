import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def create_comprehensive_excel_with_formatting(eligibility_data=None, primary_data=None, addon_data=None, AddonCoverages_data=None):
    """Create Excel file with exact format matching the reference image"""
    
    def safe_float(value):
        """Safely convert value to float, return empty string if conversion fails"""
        if value is None or value == "":
            return ""
        try:
            return float(value)
        except (ValueError, TypeError):
            return ""
    
    # Define styles
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    black_font = Font(name='Arial', size=10, bold=True, color='000000')
    header_font = Font(name='Arial', size=9, bold=True, color='000000')
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    # Create a new workbook
    wb = openpyxl.Workbook()
    # Create Eligibility sheet
    ws1 = wb.create_sheet("Eligibility")
    # Remove the default sheet
    wb.remove(wb['Sheet'])
    
    # Apply yellow background to entire header area A1:AO3
    for row in range(1, 4):  # Rows 1-3
        for col in range(1, 42):  # Columns A-AO (1-41)
            cell = ws1.cell(row=row, column=col)
            cell.fill = yellow_fill
            cell.border = thin_border
    
    # Row 1: High-level category headers
    ws1.merge_cells('A1:N1')
    ws1['A1'] = "Eligibility"
    ws1['A1'].font = black_font
    ws1['A1'].alignment = left_alignment
    ws1['A1'].number_format = '@'
    
    # Row 2: Sub-category headers
    # Relationship & Sub limit (A2:N2)
    ws1.merge_cells('A2:N2')
    ws1['A2'] = "Relationship & Sub limit"
    ws1['A2'].font = black_font
    ws1['A2'].alignment = left_alignment
    ws1['A2'].number_format = '@'
    ws1.merge_cells('O2:O2')
    ws1['O2'] = "Network List"
    ws1['O2'].font = black_font
    ws1['O2'].alignment = center_alignment
    ws1['O2'].number_format = '@'
    ws1.merge_cells('P2:P2')
    ws1['P2'] = "Provider Details"
    ws1['P2'].font = black_font
    ws1['P2'].alignment = center_alignment
    ws1['P2'].number_format = '@'
    ws1.merge_cells('Q2:Q2')
    ws1['Q2'] = "Corporate Buffer Eligibility"
    ws1['Q2'].font = black_font
    ws1['Q2'].alignment = center_alignment
    ws1['Q2'].number_format = '@'
    ws1.merge_cells('AL2:AL2')
    ws1['AL2'] = "Critical Illness Eligible"
    ws1['AL2'].font = black_font
    ws1['AL2'].alignment = center_alignment
    ws1['AL2'].number_format = '@'
    # Row 3: Detailed column headers
    column_headers_sheet3 = [
        'Max No Of Members Covered', 'Relationship Covered ', 'Relationship Covered', 'Min_Age(In Years)', 'Min_Age(In Months)',
        'Max_Age(In Years)', 'Max_Age(In Months)', 'Member_Count', 'Member_Type', 'Sublimit_Applicable',
        'Sublimit_Type', 'Sub_Limit', 'Family Buffer Applicable', 'Family Buffer Amount',
        'Is Network Applicable', 'Black listed hospitals are applicable?', 'Corporate Buffer applicable',
        'Buffer Type', 'Applicable for', 'Total Corporate Buffer', 'Corporate Buffer Limit Per Family',
        'Corporate Buffer Limit Per Parent', 'Reload of SI', 'Total Corporate Buffer',
        'Corporate Buffer Limit Per Family', 'Corporate Buffer Limit Per Parent', 'Reload of SI',
        'Approving Authority', 'Buffer OPD Limit', 'Whether increase in sum insured permissible at renewal',
        'Total Plan Buffer', 'Corporate Buffer Limit for Employee/Family', 'Corporate Buffer Limit Per Parent',
        'Reload of SI', 'Approving Authority', 'Buffer OPD Limit',
        'Whether increase in sum insured permissible at renewal', 'Critical Illness applicable',
        'Critical Illness limit per family', 'Approving Authority',
        'Whether increase in sum insured permissible at renewal'
    ]
    
    # Set the detailed headers in row 3 (A3:AO3)
    for col, header in enumerate(column_headers_sheet3, 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.alignment = left_alignment  # Left-aligned as shown in image
        cell.number_format = '@'
    
    # Set row heights for sheet 3
    ws1.row_dimensions[1].height = 18
    ws1.row_dimensions[2].height = 18
    ws1.row_dimensions[3].height = 35
    
    # Set column widths for sheet 3 - Further increased for better visibility
    column_widths_sheet3 = [
        35, 25, 40, 30, 30,  # A-E (further increased from 25,15,30,20,20)
        30, 30, 30, 30, 35,  # F-J (further increased from 20,20,20,20,25)
        30, 25, 40, 35, 30,  # K-O (further increased from 20,15,30,25,20)
        35, 35, 35, 35, 35,  # P-T (further increased from 25,25,25,25,25)
        35, 35, 35, 35, 30,  # U-Y (further increased from 25,25,25,25,20)
        35, 35, 30, 30, 30,  # Z-AD (further increased from 25,25,20,20,20)
        35, 30, 30, 35, 30,  # AE-AI (further increased from 25,20,20,25,20)
        35, 30, 30, 30, 35,  # AJ-AN (further increased from 25,20,20,20,25)
        35                    # AO (further increased from 25)
    ]
    
    for col, width in enumerate(column_widths_sheet3, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
    
    # Add actual data rows for sheet 3 (starting from row 4)
    if eligibility_data:
        for row_idx, data in enumerate(eligibility_data, 4):  # Start from row 4
            # Populate the data columns based on the extracted information
            max_members = safe_float(data.get("Max No Of Members Covered", ""))
            cell = ws1.cell(row=row_idx, column=1, value=max_members)  # A
            if max_members != "":
                cell.number_format = "#,##0"
            cell_2 = ws1.cell(row=row_idx, column=2, value=safe_float(data.get("Relationship Covered ", "")))  # B
            cell_2.number_format = "#,##0"
            cell_3 = ws1.cell(row=row_idx, column=3, value=data.get("Relationship Covered", ""))  # C
            cell_3.number_format = '@'
            cell_4 = ws1.cell(row=row_idx, column=4, value=safe_float(data.get("Min_Age(In Years)", "")))  # D
            cell_4.number_format = "#,##0"
            cell_5 = ws1.cell(row=row_idx, column=5, value=safe_float(data.get("Min_Age(In Months)", "")))  # E
            cell_5.number_format = "#,##0"
            cell_6 = ws1.cell(row=row_idx, column=6, value=safe_float(data.get("Max_Age(In Years)", "")))  # F
            cell_6.number_format = "#,##0"
            cell_7 = ws1.cell(row=row_idx, column=7, value=safe_float(data.get("Max_Age(In Months)", "")))  # G
            cell_7.number_format = "#,##0"
            cell_8 = ws1.cell(row=row_idx, column=8, value=safe_float(data.get("Member_Count", "")))  # H
            cell_8.number_format = "#,##0"
            cell_9 = ws1.cell(row=row_idx, column=9, value=data.get("Member_Type", ""))  # I
            cell_9.number_format = '@'
            cell_10 = ws1.cell(row=row_idx, column=10, value=data.get("Sublimit_Applicable", ""))  # J
            cell_10.number_format = '@'
            cell_11 = ws1.cell(row=row_idx, column=11, value=data.get("Sublimit_Type", ""))  # K
            cell_11.number_format = '@'
            cell_12 = ws1.cell(row=row_idx, column=12, value=safe_float(data.get("Sub_Limit", "")))  # L
            cell_12.number_format = "#,##0"
            cell_13 = ws1.cell(row=row_idx, column=13, value=data.get("Family Buffer Applicable", ""))  # M
            cell_13.number_format = '@'
            cell_14 = ws1.cell(row=row_idx, column=14, value=safe_float(data.get("Family Buffer Amount", "")))  # N
            cell_14.number_format = "#,##0"
            cell_15 = ws1.cell(row=row_idx, column=15, value=data.get("Is Network Applicable", ""))  # O
            cell_15.number_format = '@'
            cell_16 = ws1.cell(row=row_idx, column=16, value=data.get("Black listed hospitals are applicable?", ""))  # P
            cell_16.number_format = '@'
            cell_17 = ws1.cell(row=row_idx, column=17, value=data.get("Corporate Buffer applicable", ""))  # O
            cell_17.number_format = '@'
            cell_18 = ws1.cell(row=row_idx, column=18, value=data.get("Buffer Type", ""))  # P
            cell_18.number_format = '@'
            cell_19 = ws1.cell(row=row_idx, column=19, value=data.get("Applicable for", ""))  # Q
            cell_19.number_format = '@'
            cell_20 = ws1.cell(row=row_idx, column=20, value=safe_float(data.get("Total Corporate Buffer", "")))  # R
            cell_20.number_format = "#,##0"
            cell_21 = ws1.cell(row=row_idx, column=21, value=safe_float(data.get("Corporate Buffer Limit Per Family", "")))  # S
            cell_21.number_format = "#,##0"
            cell_22 = ws1.cell(row=row_idx, column=22, value=safe_float(data.get("Corporate Buffer Limit Per Parent", "")))  # T
            cell_22.number_format = "#,##0"
            ws1.cell(row=row_idx, column=23, value=data.get("Reload of SI", ""))  # U
            
            cell_24 = ws1.cell(row=row_idx, column=24, value=safe_float(data.get("Total Corporate Buffer.1", "")))  # V
            cell_24.number_format = "#,##0"
            cell_25 = ws1.cell(row=row_idx, column=25, value=safe_float(data.get("Corporate Buffer Limit Per Family.1", "")))  # W
            cell_25.number_format = "#,##0"
            cell_26 = ws1.cell(row=row_idx, column=26, value=safe_float(data.get("Corporate Buffer Limit Per Parent.1", "")))  # X
            cell_26.number_format = "#,##0"
            cell_27 = ws1.cell(row=row_idx, column=27, value=data.get("Reload of SI.1", ""))
            cell_27.number_format = '@'
            cell_28 = ws1.cell(row=row_idx, column=28, value=data.get("Approving Authority", ""))
            cell_28.number_format = '@'
            cell_29 = ws1.cell(row=row_idx, column=29, value=data.get("Buffer OPD Limit", ""))
            cell_29.number_format = '@'  
            ws1.cell(row=row_idx, column=30, value=data.get("Whether increase in sum insured permissible at renewal", "")).number_format = '@'  
            ws1.cell(row=row_idx, column=31, value=safe_float(data.get("Total Plan Buffer", ""))).number_format = "#,##0"  
            ws1.cell(row=row_idx, column=32, value=safe_float(data.get("Corporate Bufferr Limit for Employee/Family", ""))).number_format = "#,##0"  
            ws1.cell(row=row_idx, column=33, value=safe_float(data.get("Corporate Buffer Limit Per Parent.2", ""))).number_format = "#,##0"  
            ws1.cell(row=row_idx, column=34, value=data.get("Reload of SI.2", "")).number_format = '@'  
            ws1.cell(row=row_idx, column=35, value=data.get("Approving Authority.1", "")).number_format = '@'  
            ws1.cell(row=row_idx, column=36, value=data.get("Buffer OPD Limit.1", "")).number_format = '@'  
            ws1.cell(row=row_idx, column=37, value=data.get("Whether increase in sum insured permissible at renewal.1", "")).number_format = '@'  
            
            # Critical Illness fields
            ws1.cell(row=row_idx, column=38, value=data.get("Critical Illness applicable", "")).number_format = '@'  
            ws1.cell(row=row_idx, column=39, value=safe_float(data.get("Critical Illness limit per family", ""))).number_format = "#,##0.00"  
            ws1.cell(row=row_idx, column=40, value=data.get("Critical Illness Approving Authority", "")).number_format = '@'  
            ws1.cell(row=row_idx, column=41, value=data.get("Critical Illness Whether increase in sum insured permissible at renewal", "")).number_format = '@'  
            
            # Apply borders to all data cells
            for col in range(1, 42):  # A-AO (1-41)
                cell = ws1.cell(row=row_idx, column=col)
                cell.border = thin_border
    
    # Add empty rows with borders for remaining rows
    start_row = 4 + (len(eligibility_data) if eligibility_data else 0)
    for row in range(start_row, max(10, start_row + 2)):  # Add at least a few empty rows
        for col in range(1, 42):  # A-AO (1-41)
            cell = ws1.cell(row=row, column=col)
            cell.border = thin_border

    # Create fifth sheet: Primary Cover
    ws2 = wb.create_sheet("Primary Cover")

    # Apply yellow background to rows 1-3 and borders to rows 1-4
    for row in range(1, 5):  # Rows 1-4
        for col in range(1, 85):  # Columns A-CG (1-83)
            cell = ws2.cell(row=row, column=col)
            
            # Apply yellow background only to rows 1-3
            if row <= 3:
                cell.fill = yellow_fill
            
            # Apply border to all rows 1-4
            cell.border = thin_border

    # Row 1: High-level category headers
    # Primary Cover (A1:C1)
    ws2.merge_cells('A1:C1')
    ws2['A1'] = "Primary Cover"
    ws2['A1'].font = black_font
    ws2['A1'].alignment = left_alignment

    # Row 2: Sub-category headers
    # Pre & Post Hospitalization (I2:S2)
    ws2.merge_cells('A2:S2')
    ws2['A2'] = "Pre & Post Hospitalization"
    ws2['A2'].font = black_font
    ws2['A2'].alignment = center_alignment

    # Maternity (T2:U2)
    ws2.merge_cells('U2:U2')
    ws2['U2'] = "Maternity"
    ws2['U2'].font = black_font
    ws2['U2'].alignment = center_alignment

    

    ws2.merge_cells('AI2:AI2')
    ws2['AI2'] = "Normal"
    ws2['AI2'].font = black_font
    ws2['AI2'].alignment = center_alignment

    ws2.merge_cells('AP2:AP2')
    ws2['AP2'] = "Caesarian"
    ws2['AP2'].font = black_font
    ws2['AP2'].alignment = center_alignment

    ws2.merge_cells('AW2:AW2')
    ws2['AW2'] = "Critical"
    ws2['AW2'].font = black_font
    ws2['AW2'].alignment = center_alignment

    

    # Row 3: Individual column headers
    column_headers_row3 = [
    'Benefit Applicable?', 'Is Pre and Post Combined?', 'Type Of Expense', 'No. Of Days', '% Limit Applicable On',
    '% Limit', 'Limit', 'Applicability', 'Type of expense', 'No. Of Days', '% Limit Applicable',
    'Limit Percentage', 'Limit Amount', 'Applicability', 'Type of expense', 'No. Of Days',
    '% Limit Applicable', 'Limit Percentage', 'Limit Amount', 'Applicability', 'Benefit Applicable?',
    'Waiting Period(In Days)', 'Limit On Number Of Live Children', 'Member Contribution Applicable?',
    'Copay or deductible Applicable?', 'Is Maternity Combined?', 'Sum Insured', '% Limit', 'Limit','Limit',
    'Applicability', 'Copay', 'Deductible', 'Is Maternity Combined?', 'Sum Insured', '% Limit',
    'Limit', 'Limit', 'Applicability', 'Copay', 'Deductible', 'Sum Insured', '% Limit', 'Limit', 'Limit',
    'Applicability', 'Copay', 'Deductible', 'Sum Insured', '% Limit', 'Limit', 'Limit', 'Applicability',
    'Copay', 'Deductible', 'Pre&Post Natal Applicable', 'Over & Above Maternity Limit',
    'Is Pre&Post Natal Combined?', 'Maternity', 'No. Of Days', '% Limit Applicable On', '% Limit',
    'Limit', 'Applicability', 'Maternity', 'No. Of Days', '% Limit Applicable On', '% Limit', 'Limit',
    'Applicability', 'Maternity', 'No. Of Days', '% Limit Applicable On', '% Limit', 'Limit',
    'Applicability', 'New Born Covered?', 'Covered From', 'Is New Born Limit Applicable', 'Sum Insured',
    '% Limit Applicable On', 'Limit Percentage', 'Limit Amount', 'Applicability'
]



    for col, header in enumerate(column_headers_row3, 1):  # Start from column C (3)
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = black_font
        cell.alignment = left_alignment

    if primary_data:
        for row_idx, data in enumerate(primary_data, 4):
            # Combined section (columns 1-8)
            ws2.cell(row=row_idx, column=1, value=data.get("Combined_Benefit_Applicable", "Yes")).number_format = '@'
            ws2.cell(row=row_idx, column=2, value=data.get("Combined_Is_Pre_and_Post_Combined", "")).number_format = '@'
            ws2.cell(row=row_idx, column=3, value=data.get("Combined_Type_Of_Expense", "")).number_format = '@'
            ws2.cell(row=row_idx, column=4, value=data.get("Combined_No_Of_Days", ""))
            ws2.cell(row=row_idx, column=5, value=data.get("Combined_Percent_Limit_Applicable_On", "")).number_format = '@'
            ws2.cell(row=row_idx, column=6, value=data.get("Combined_Percent_Limit", "")).number_format = "0.00%"
            ws2.cell(row=row_idx, column=7, value=data.get("Combined_Limit", ""))
            ws2.cell(row=row_idx, column=8, value=data.get("Combined_Applicability", "")).number_format = '@'
            
            # Pre Hospitalisation section (columns 9-14)
            ws2.cell(row=row_idx, column=9, value=data.get("Type of expense 1", "")).number_format = '@'
            ws2.cell(row=row_idx, column=10, value=data.get("No. Of Days 1", "")).number_format = '#,##0'
            ws2.cell(row=row_idx, column=11, value=data.get("% Limit Applicable 1", "")).number_format = '@'
            ws2.cell(row=row_idx, column=12, value=data.get("Limit Percentage 1", "")).number_format = "0.00%"
            ws2.cell(row=row_idx, column=13, value=data.get("Limit Amount_19", "500000")).number_format = "#,##0"
            ws2.cell(row=row_idx, column=14, value=data.get("Applicability 1", "")).number_format = '@'
            
            # Post Hospitalisation/OPD section (columns 15-20)
            ws2.cell(row=row_idx, column=15, value=data.get("Type of expense 2", "")).number_format = '@'
            ws2.cell(row=row_idx, column=16, value=data.get("No. Of Days 2", "")).number_format = '#,##0'
            ws2.cell(row=row_idx, column=17, value=data.get("% Limit Applicable 2", "")).number_format = '@'
            ws2.cell(row=row_idx, column=18, value=data.get("Limit Percentage 2", "")).number_format = "0.00%"
            ws2.cell(row=row_idx, column=19, value=data.get("Limit Amount_20", "500000")).number_format = "#,##0"
            ws2.cell(row=row_idx, column=20, value=data.get("Applicability 2", "")).number_format = '@'

                        # Main Maternity section (columns 21-32)
            ws2.cell(row=row_idx, column=21, value=data.get("Maternity Benefit Applicable?", "No"))
            
            # Maternity Benefit Applicable Logic
            if data.get("Maternity Benefit Applicable?", "").lower() == "no":
                # If No: Return empty for columns 22-75
                for col in range(22, 76):
                    ws2.cell(row=row_idx, column=col, value="")
            else:
                # If Yes: Continue with all maternity logic
                ws2.cell(row=row_idx, column=22, value=data.get("Maternity Waiting Period(In Days)", "")).number_format = '#,##0'
                ws2.cell(row=row_idx, column=23, value=data.get("Maternity Limit On Number Of Live Children", "")).number_format = '#,##0'
                ws2.cell(row=row_idx, column=24, value=data.get("Maternity Member Contribution Applicable?", "")).number_format = '@'
                ws2.cell(row=row_idx, column=25, value=data.get("Maternity Copay or deductible Applicable?", "")).number_format = '@'
                ws2.cell(row=row_idx, column=26, value=data.get("Maternity Is Combined?", "")).number_format = '@'
                
                # Maternity Is Combined? Logic
                if data.get("Maternity Is Combined?", "") == "No":
                    # If No: Return empty for columns 27-32, return original data for columns 34-48
                    ws2.cell(row=row_idx, column=27, value="")
                    ws2.cell(row=row_idx, column=28, value="")
                    ws2.cell(row=row_idx, column=29, value="")
                    ws2.cell(row=row_idx, column=30, value="")
                    ws2.cell(row=row_idx, column=31, value="")
                    ws2.cell(row=row_idx, column=32, value="")
                    ws2.cell(row=row_idx, column=33, value="")
                    
                    # Return original data for columns 34-48
                    ws2.cell(row=row_idx, column=34, value=data.get("Normal_Sum_Insured", "500000")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=35, value=data.get("Normal_Limit_%", "Sum Insured")).number_format = "0.00%"
                    ws2.cell(row=row_idx, column=36, value=data.get("Normal Delivery Limit", "")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=37, value=data.get("Normal Delivery Limit", "")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=38, value=data.get("Normal_Applicability", "Lower")).number_format = '@'
                    ws2.cell(row=row_idx, column=39, value=data.get("Normal_copay", "")).number_format = '#,##0'
                    ws2.cell(row=row_idx, column=40, value=data.get("Normal_Delivery_Applicability", "")).number_format = '@'
                    ws2.cell(row=row_idx, column=41, value=data.get("Ceaserean_sum_insured", "500000")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=42, value=data.get("Ceaserean_Limit_%", "Sum Insured")).number_format = "0.00%"
                    ws2.cell(row=row_idx, column=43, value=data.get("Caesarean Limit", "")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=44, value=data.get("Caesarean Limit", "")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=45, value=data.get("Caesarean_applicability", "Lower")).number_format = '@'
                    ws2.cell(row=row_idx, column=46, value=data.get("Ceaserean_copay", "")).number_format = '#,##0'
                    ws2.cell(row=row_idx, column=47, value=data.get("Caesarean_Applicability", "")).number_format = '@'
                    ws2.cell(row=row_idx, column=48, value=data.get("critical_sum_insured", "")).number_format = "#,##0"
                    
                elif data.get("Maternity Is Combined?", "") == "Yes":
                    # If Yes: Return original data for columns 27-32, return empty for columns 34-48
                    ws2.cell(row=row_idx, column=27, value=data.get("Maternity Sum Insured_1", "500000")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=28, value=data.get("Maternity % Limit_1", "Sum Insured")).number_format = "0.00%"
                    ws2.cell(row=row_idx, column=29, value=data.get("Normal Delivery Limit", "")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=30, value=data.get("Normal Delivery Limit", "")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=31, value=data.get("Maternity Applicability_10", "Lower")).number_format = '@'
                    ws2.cell(row=row_idx, column=32, value=data.get("Maternity Copay_1", "")).number_format = '#,##0'
                    ws2.cell(row=row_idx, column=33, value=data.get("Maternity Deductible_1", "")).number_format = '#,##0'
                    ws2.cell(row=row_idx, column=34, value=data.get("Maternity Is Combined?", "")).number_format = '@'
                    
                    # Return empty for columns 34-48
                    for col in range(34, 49):
                        ws2.cell(row=row_idx, column=col, value="")
                        
                else:
                    # Default case: Use original logic
                    ws2.cell(row=row_idx, column=27, value=data.get("Maternity Sum Insured", ""))
                    ws2.cell(row=row_idx, column=28, value=data.get("Maternity % Limit", ""))
                    ws2.cell(row=row_idx, column=29, value=data.get("Maternity_Limit_amount", ""))
                    ws2.cell(row=row_idx, column=30, value=data.get("Maternity_Limit_amount", ""))
                    ws2.cell(row=row_idx, column=31, value=data.get("Maternity Applicability", ""))
                    ws2.cell(row=row_idx, column=32, value=data.get("Maternity Copay", ""))
                    ws2.cell(row=row_idx, column=33, value=data.get("Maternity Deductible", ""))
                    ws2.cell(row=row_idx, column=34, value=data.get("Maternity Is Combined?", ""))

                # Pre-Natal section (columns 34-48) - Only populate if Maternity Is Combined is NOT "Yes"
                if data.get("Maternity Is Combined?", "") != "Yes":
                    ws2.cell(row=row_idx, column=34, value=data.get("Maternity Is Combined?", "")).number_format = '@'
                    ws2.cell(row=row_idx, column=35, value=data.get("Normal_Sum_Insured", "500000")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=36, value=data.get("Normal_Limit_%", "Sum Insured")).number_format = "0.00%"
                    ws2.cell(row=row_idx, column=37, value=data.get("Normal Delivery Limit", "")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=38, value=data.get("Normal Delivery Limit", "")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=39, value=data.get("Normal_Applicability", "Lower")).number_format = '@'
                    ws2.cell(row=row_idx, column=40, value=data.get("Normal_copay", "")).number_format = '#,##0'
                    ws2.cell(row=row_idx, column=41, value=data.get("Normal_Delivery_Applicability", "")).number_format = '@'
                    ws2.cell(row=row_idx, column=42, value=data.get("Ceaserean_sum_insured", "500000")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=43, value=data.get("Ceaserean_Limit_%", "Sum Insured")).number_format = "0.00%"
                    ws2.cell(row=row_idx, column=44, value=data.get("Caesarean Limit", "")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=45, value=data.get("Caesarean Limit", "")).number_format = "#,##0"
                    ws2.cell(row=row_idx, column=46, value=data.get("Caesarean_applicability", "Lower")).number_format = '@'
                    ws2.cell(row=row_idx, column=47, value=data.get("Ceaserean_copay", "")).number_format = '#,##0'
                    ws2.cell(row=row_idx, column=48, value=data.get("Caesarean_Applicability", "")).number_format = '@'
                    ws2.cell(row=row_idx, column=49, value=data.get("critical_sum_insured", "")).number_format = "#,##0"
                else:
                    # Leave columns 34-48 empty if Maternity Is Combined is "Yes"
                    for col in range(35, 50):
                        ws2.cell(row=row_idx, column=col, value="")

                # Post-Natal section (columns 49-61)
                ws2.cell(row=row_idx, column=50, value=data.get("Post-Natal Deductible", ""))
                ws2.cell(row=row_idx, column=51, value=data.get("critical_Limit", ""))
                ws2.cell(row=row_idx, column=52, value=data.get("critical_Limit", ""))
                ws2.cell(row=row_idx, column=53, value=data.get("critical_Limit", ""))
                ws2.cell(row=row_idx, column=54, value=data.get("critical_Limit", ""))
                ws2.cell(row=row_idx, column=55, value=data.get("Critical_Applicability", ""))
                ws2.cell(row=row_idx, column=56, value=data.get("Pre-Natal and Post-Natal Expenses Covered", "No")).number_format = '@'
                ws2.cell(row=row_idx, column=57, value=data.get("Over-Above-Maternity Limit Applicable", "No")).number_format = '@'
                ws2.cell(row=row_idx, column=58, value=data.get("Is Pre&Post Natal Combined?", "")).number_format = '@'
                
                # Pre&Post Natal Combined Logic
                if data.get("Is Pre&Post Natal Combined?", "") == "No":
                    # If No: Return empty for columns 58-63, return original data for columns 64-75
                    ws2.cell(row=row_idx, column=59, value="")
                    ws2.cell(row=row_idx, column=60, value="")
                    ws2.cell(row=row_idx, column=61, value="")
                    ws2.cell(row=row_idx, column=62, value="")
                    ws2.cell(row=row_idx, column=63, value="")
                    ws2.cell(row=row_idx, column=64, value="")
                    
                    # Return original data for columns 64-75
                    ws2.cell(row=row_idx, column=65, value=data.get("Maternity_hardcode_2", "Pre-Natal Expenses")).number_format = '@'
                    ws2.cell(row=row_idx, column=66, value=data.get("No.of Days_6", "30")).number_format = '#,##0'
                    ws2.cell(row=row_idx, column=67, value=data.get("Sum Insured_3", "Sum Insured")).number_format = '@'
                    ws2.cell(row=row_idx, column=68, value=data.get("Pre-Natal Limit Calc Percentage", "")).number_format = '@'
                    ws2.cell(row=row_idx, column=69, value=data.get("Pre-Natal Limit", "")).number_format = '#,##0'
                    ws2.cell(row=row_idx, column=70, value=data.get("Aplicability_pre", "Lower")).number_format = '@'
                    ws2.cell(row=row_idx, column=71, value=data.get("Maternity_hardcode_3", "Post-Natal Expenses")).number_format = '@'
                    ws2.cell(row=row_idx, column=72, value=data.get("No.of Days_10", "60")).number_format = '#,##0'
                    ws2.cell(row=row_idx, column=73, value=data.get("sum insured_11", "Sum Insured")).number_format = '@'
                    ws2.cell(row=row_idx, column=74, value=data.get("Pre-Natal Limit Calc Percentage", "")).number_format = '@'
                    ws2.cell(row=row_idx, column=75, value=data.get("Post-Natal Limit", "")).number_format = '#,##0'
                    ws2.cell(row=row_idx, column=76, value=data.get("Applicability_post", "Lower")).number_format = '@'
                    
                elif data.get("Is Pre&Post Natal Combined?", "") == "Yes":
                    # If Yes: Return original data for columns 58-63, return empty for columns 64-75
                    ws2.cell(row=row_idx, column=59, value=data.get("Materninity_hardcode", "Pre & Post Natal Expenses")).number_format = '@'
                    ws2.cell(row=row_idx, column=60, value=data.get("No.of Days_5", "30")).number_format = '#,##0'
                    ws2.cell(row=row_idx, column=61, value=data.get("Sum Insured_2", "Sum Insured")).number_format = '@'
                    ws2.cell(row=row_idx, column=62, value=data.get("Pre-Natal Limit Calc Percentage", "")).number_format = '@'
                    ws2.cell(row=row_idx, column=63, value=data.get("Pre-Natal Limit", "")).number_format = '#,##0'
                    ws2.cell(row=row_idx, column=64, value=data.get("Applicability_7", "Lower")).number_format = '@'
                    
                    # Return empty for columns 64-75
                    for col in range(65, 77):
                        ws2.cell(row=row_idx, column=col, value="")
                        
                else:
                    # Default case: Use original logic
                    ws2.cell(row=row_idx, column=58, value=data.get("Materninity_hardcode", "Pre & Post Natal Expenses"))
                    ws2.cell(row=row_idx, column=59, value=data.get("No.of Days_5", "30"))
                    ws2.cell(row=row_idx, column=60, value=data.get("Sum Insured_2", "Sum Insured"))
                    ws2.cell(row=row_idx, column=61, value=data.get("Pre-Natal Limit Calc Percentage", ""))
                    ws2.cell(row=row_idx, column=62, value=data.get("Pre-Natal Limit", ""))
                    ws2.cell(row=row_idx, column=63, value=data.get("Applicability_7", "Lower"))
                    ws2.cell(row=row_idx, column=64, value=data.get("Maternity_hardcode_2", "Pre-Natal Expenses"))
                    ws2.cell(row=row_idx, column=65, value=data.get("No.of Days_6", ""))
                    ws2.cell(row=row_idx, column=66, value=data.get("Sum Insured_3", ""))
                    ws2.cell(row=row_idx, column=67, value=data.get("Limit_8", ""))
                    ws2.cell(row=row_idx, column=68, value=data.get("Post-Natal Limit", ""))
                    ws2.cell(row=row_idx, column=69, value=data.get("%Limit", ""))
                    ws2.cell(row=row_idx, column=70, value=data.get("Limit", ""))
                    ws2.cell(row=row_idx, column=71, value=data.get("applicability", ""))
                    ws2.cell(row=row_idx, column=72, value=data.get("maternity_2", ""))
                    ws2.cell(row=row_idx, column=73, value=data.get("No.of Days_2", ""))
                    ws2.cell(row=row_idx, column=74, value=data.get("%Limit Applicable on_3", ""))
                    ws2.cell(row=row_idx, column=75, value=data.get("%limit_2", ""))
            #New Born covered logic 
            newborn_covered = data.get("new born covered?", "")
            # Set column 76 to "No" if not "yes", otherwise use the actual value
            if newborn_covered.lower() != "yes":
                ws2.cell(row=row_idx, column=77, value="No")
            else:
                ws2.cell(row=row_idx, column=77, value=newborn_covered)
            
            # Only populate columns 77-85 if "new born covered?" is "yes"
            if newborn_covered.lower() == "yes":
                ws2.cell(row=row_idx, column=78, value=data.get("covered From", ""))

                # New Born section (columns 78-85) - Using maternity_26_columns.py logic
                ws2.cell(row=row_idx, column=79, value=data.get("Is New Born Limit Applicable", "")).number_format = '@'
                ws2.cell(row=row_idx, column=80, value=data.get("covered From_1", "")).number_format = '@'
                ws2.cell(row=row_idx, column=81, value=data.get("Is New Born Limit Applicable_1", "")).number_format = '@'
                ws2.cell(row=row_idx, column=82, value=data.get("Newborn_sum_insured", "")).number_format = '#,##0'
                ws2.cell(row=row_idx, column=83, value=data.get("Newborn_Limit_applicable_on", "")).number_format = '@'
                ws2.cell(row=row_idx, column=84, value=data.get("Newborn_Limit_percentage", "")).number_format = '@'
                ws2.cell(row=row_idx, column=85, value=data.get("Newborn_Limit_amount", "")).number_format = '#,##0'
                ws2.cell(row=row_idx, column=86, value=data.get("Newborn_applicability", "")).number_format = '@'
            else:
                # Leave columns 77-85 empty if "new born covered?" is not "yes"
                for col in range(78, 87):
                    ws2.cell(row=row_idx, column=col, value="")

            # Apply borders to all data cells
            for col in range(1, 85):  # A-CG
                cell = ws2.cell(row=row_idx, column=col)
                cell.border = thin_border

    

    # Set row heights for sheet 5
    ws2.row_dimensions[1].height = 18
    ws2.row_dimensions[2].height = 18
    ws2.row_dimensions[3].height = 18
    ws2.row_dimensions[4].height = 35

    # Set column widths for sheet 5 - Further increased for better visibility
    column_widths_sheet5 = [
        35, 35, 45, 30, 35,  # A-E (further increased from 25,25,35,20,25)
        30, 30, 35,  # F-H (further increased from 20,20,25)
        40, 30, 35, 35, 35, 35,  # I-N (further increased from 30,20,25,25,25,25)
        40, 30, 35, 35, 35, 35,  # O-T (further increased from 30,20,25,25,25,25)
        30  # U (further increased from 20)
    ]

    # Extend widths till CG (83 columns) with default width 35 (further increased from 25)
    while len(column_widths_sheet5) < 83:
        column_widths_sheet5.append(35)

    for col, width in enumerate(column_widths_sheet5, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Add data rows for sheet 5 (starting from row 5) with borders only
    for row in range(5, 6):  # Rows 5-26 as shown in image (fixed range)
        for col in range(1, 84):  # A-CG
            cell = ws2.cell(row=row, column=col)
            cell.border = thin_border
            
    # Create sixth sheet: Secondary Cover
    ws3 = wb.create_sheet("Addon covers")

    # Apply yellow background to rows 1-3 and borders to rows 1-4
    for row in range(1, 5):  # Rows 1-4
        for col in range(1, 33):  
            cell = ws3.cell(row=row, column=col)
            # Apply yellow background only to rows 1-3
            if row <= 3:
                cell.fill = yellow_fill
            
            # Apply border to all rows 1-4
            cell.border = thin_border

    # Row 1: High-level category headers
    ws3.merge_cells('A1:AF1')
    ws3['A1'] = "Addon Covers"
    ws3['A1'].font = black_font
    ws3['A1'].alignment = left_alignment
    ws3['A1'].number_format = '@'

    # Row 2: Sub-category headers
    ws3.merge_cells('A2:AF2')
    ws3['A2'] = ""
    ws3['A2'].font = black_font
    ws3['A2'].alignment = center_alignment
    ws3['A2'].number_format = '@'
    column_headers_row3 = ["Ambulance Cover", "Anyone Illness", "Attendant Care", "Cancer Cover", "Convalescence Benefit",
    "Critical Illness Benefit", "Daily/Hospital Cash Benefit", "Dental Cover", "Diabetic Cover",
    "Doctor & Nurse Home Visit Cover", "Education Fund", "Funeral", "Getwell Benefit",
    "Hardship Critical Illness Cover", "Health Check up", "Hypertension Cover",
    "Intensive Care Benefit", "Loss Of Pay Cover", "Medical Evacuation Cover",
    "Medical Second Opinion", "Non Medical Expense Cover", "Out Patient Cover", "Optical Cover",
    "Organ Donor Medical Expense Cover", "Personal Accident Cover", "Pre Existing Disease Benefit",
    "Psychiatric Cover", "Recovery Benefit", "Referral Hospital Care", "Surgical Benefit",
    "Top Up Cover", "Vaccination/Immunization Cover"
]

    for col, header in enumerate(column_headers_row3, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = black_font
        cell.alignment = left_alignment
        cell.number_format = '@'
        
    if addon_data:
        for row_idx, data in enumerate(addon_data, 4):
             ws3.cell(row=row_idx, column=1, value=data.get("Ambulance Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=2, value=data.get("Anyone Illness", "")).number_format = '@'
             ws3.cell(row=row_idx, column=3, value=data.get("Attendant Care", "")).number_format = '@'
             ws3.cell(row=row_idx, column=4, value=data.get("Cancer Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=5, value=data.get("Convalescence Benefit", "")).number_format = '@'
             ws3.cell(row=row_idx, column=6, value=data.get("Critical Illness Benefit", "")).number_format = '@'
             ws3.cell(row=row_idx, column=7, value=data.get("Daily/Hospital Cash Benefit", "")).number_format = '@'
             ws3.cell(row=row_idx, column=8, value=data.get("Dental Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=9, value=data.get("Diabetic Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=10, value=data.get("Doctor & Nurse Home Visit Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=11, value=data.get("Education Fund", "")).number_format = '@'
             ws3.cell(row=row_idx, column=12, value=data.get("Funeral", "")).number_format = '@'
             ws3.cell(row=row_idx, column=13, value=data.get("Getwell Benefit", "")).number_format = '@'
             ws3.cell(row=row_idx, column=14, value=data.get("Hardship Critical Illness Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=15, value=data.get("Health Check up", "")).number_format = '@'
             ws3.cell(row=row_idx, column=16, value=data.get("Hypertension Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=17, value=data.get("Intensive Care Benefit", "")).number_format = '@'
             ws3.cell(row=row_idx, column=18, value=data.get("Loss Of Pay Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=19, value=data.get("Medical Evacuation Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=20, value=data.get("Medical Second Opinion", "")).number_format = '@'
             ws3.cell(row=row_idx, column=21, value=data.get("Non Medical Expense Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=22, value=data.get("Out Patient Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=23, value=data.get("Optical Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=24, value=data.get("Organ Donor Medical Expense Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=25, value=data.get("Personal Accident Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=26, value=data.get("Pre Existing Disease Benefit", "")).number_format = '@'
             ws3.cell(row=row_idx, column=27, value=data.get("Psychiatric Cover", "")).number_format = '@'  
             ws3.cell(row=row_idx, column=28, value=data.get("Recovery Benefit", "")).number_format = '@'
             ws3.cell(row=row_idx, column=29, value=data.get("Referral Hospital Care", "")).number_format = '@'
             ws3.cell(row=row_idx, column=30, value=data.get("Surgical Benefit", "")).number_format = '@'
             ws3.cell(row=row_idx, column=31, value=data.get("Top Up Cover", "")).number_format = '@'
             ws3.cell(row=row_idx, column=32, value=data.get("Vaccination/Immunization Cover", "")).number_format = '@'
                

    # Set row heights for sheet 6
    ws3.row_dimensions[1].height = 18
    ws3.row_dimensions[2].height = 18
    ws3.row_dimensions[3].height = 35

    # Set column widths for sheet 6 - Further increased for better visibility
    column_widths_sheet6 = [
        35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35, 35
    ]
    
    for col, width in enumerate(column_widths_sheet6, 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width  
    
    # Add data rows for sheet 6 (starting from row 4) with borders only
    for row in range(4, 5):  # Rows 4-27 as shown in image
        for col in range(1, 32):  # A-AF (1-32)
            cell = ws3.cell(row=row, column=col)
            cell.border = thin_border

    # Create seventh sheet: Addon Coverages
    ws4 = wb.create_sheet("Addon Coverages")

    # Apply yellow background to rows 1-3 and borders to rows 1-4
    for row in range(1, 5):  # Rows 1-4
        for col in range(1, 449):  # A to QF (1-448, so range needs to be 1-449)
            cell = ws4.cell(row=row, column=col)
            # Apply yellow background only to rows 1-3
            if row <= 3:
                cell.fill = yellow_fill
            
            # Apply border to all rows 1-4
            cell.border = thin_border

    ws4.merge_cells('A1:QF1')
    ws4['A1'] = "Addon Coverages" 
    ws4['A1'].font = black_font
    ws4['A1'].alignment = center_alignment

    # Row 2: High-level category headers
    ws4.merge_cells('A2:F2')
    ws4['A2'] = "Ambulance Cover"
    ws4['A2'].font = black_font
    ws4['A2'].alignment = center_alignment
    ws4['A2'].number_format = '@'
    ws4.merge_cells('G2:Y2')
    ws4['G2'] = "Anyone Illness"
    ws4['G2'].font = black_font
    ws4['G2'].alignment = center_alignment
    ws4['G2'].number_format = '@'
    ws4.merge_cells('Z2:AQ2')
    ws4['Z2'] = "Attendant Care"
    ws4['Z2'].font = black_font
    ws4['Z2'].alignment = center_alignment
    ws4['Z2'].number_format = '@'
    ws4.merge_cells('AR2:AR2')
    ws4['AR2'] = "Cancer Cover"
    ws4['AR2'].font = black_font
    ws4['AR2'].alignment = center_alignment
    ws4['AR2'].number_format = '@'
    ws4.merge_cells('AS2:AV2')
    ws4['AS2'] = "Convalescence Benefit"
    ws4['AS2'].font = black_font
    ws4['AS2'].alignment = center_alignment
    ws4['AS2'].number_format = '@'
    ws4.merge_cells('AW2:BO2')
    ws4['AW2'] = "Critical Illness"
    ws4['AW2'].font = black_font
    ws4['AW2'].alignment = center_alignment
    ws4['AW2'].number_format = '@'
    ws4.merge_cells('BP2:CD2')
    ws4['BP2'] = "Daily Cash"
    ws4['BP2'].font = black_font
    ws4['BP2'].alignment = left_alignment
    ws4['BP2'].number_format = '@'
    ws4.merge_cells('CE2:CY2')
    ws4['CE2'] = "Dental Cover"
    ws4['CE2'].font = black_font
    ws4['CE2'].alignment = center_alignment
    ws4['CE2'].number_format = '@'
    ws4.merge_cells('CZ2:CZ2')
    ws4['CZ2'] = "Diabetic Cover"
    ws4['CZ2'].font = black_font
    ws4['CZ2'].alignment = center_alignment
    ws4['CZ2'].number_format = '@'
    ws4.merge_cells('DA2:DG2')
    ws4['DA2'] = "Applicability of Doctor's Home Visit & Nursing Charges"
    ws4['DA2'].font = black_font
    ws4['DA2'].alignment = left_alignment
    ws4['DA2'].number_format = '@'
    ws4.merge_cells('DH2:DN2')
    ws4['DH2'] = "Education Fund"
    ws4['DH2'].font = black_font
    ws4['DH2'].alignment = center_alignment
    ws4['DH2'].number_format = '@'  
    ws4.merge_cells('DO2:ED2')
    ws4['DO2'] = "Funeral Expenses"
    ws4['DO2'].font = black_font
    ws4['DO2'].alignment = center_alignment
    ws4['DO2'].number_format = '@'
    ws4.merge_cells('EE2:ET2')
    ws4['EE2'] = "Get Well Benefit"
    ws4['EE2'].font = black_font
    ws4['EE2'].alignment = center_alignment
    ws4['EE2'].number_format = '@'
    ws4.merge_cells('EU2:FO2')
    ws4['EU2'] = "Hardship Critical Illness Cover"
    ws4['EU2'].font = black_font
    ws4['EU2'].alignment = center_alignment
    ws4['EU2'].number_format = '@'
    ws4.merge_cells('FP2:GH2')
    ws4['FP2'] = "Health Check-up"
    ws4['FP2'].font = black_font
    ws4['FP2'].alignment = center_alignment
    ws4['FP2'].number_format = '@'
    ws4.merge_cells('GI2:GI2')
    ws4['GI2'] = "Hypertension Cover"
    ws4['GI2'].font = black_font
    ws4['GI2'].alignment = center_alignment
    ws4['GI2'].number_format = '@'
    # Fixed the overlapping merge - this was causing issues
    ws4.merge_cells('GJ2:GQ2')
    ws4['GJ2'] = "Intensive Care Benefit"
    ws4['GJ2'].font = black_font
    ws4['GJ2'].alignment = center_alignment
    ws4['GJ2'].number_format = '@'
    ws4.merge_cells('GR2:HH2')
    ws4['GR2'] = "Loss Of Pay"
    ws4['GR2'].font = black_font
    ws4['GR2'].alignment = center_alignment
    ws4['GR2'].number_format = '@'
    ws4.merge_cells('HI2:HO2')
    ws4['HI2'] = "Medical Evacuation"
    ws4['HI2'].font = black_font
    ws4['HI2'].alignment = center_alignment
    ws4['HI2'].number_format = '@'
    ws4.merge_cells('HP2:IF2')
    ws4['HP2'] = "Medical Second Opinion"
    ws4['HP2'].font = black_font
    ws4['HP2'].alignment = center_alignment
    ws4['HP2'].number_format = '@'
    ws4.merge_cells('IG2:IW2')
    ws4['IG2'] = "Non Medical Expense"
    ws4['IG2'].font = black_font
    ws4['IG2'].alignment = center_alignment
    ws4['IG2'].number_format = '@'
    ws4.merge_cells('IX2:KN2')
    ws4['IX2'] = "Out Patient Configuration"
    ws4['IX2'].font = black_font
    ws4['IX2'].alignment = center_alignment
    ws4['IX2'].number_format = '@'
    ws4.merge_cells('KO2:LO2')
    ws4['KO2'] = "Optical Cover"
    ws4['KO2'].font = black_font
    ws4['KO2'].alignment = center_alignment
    ws4['KO2'].number_format = '@'      
    # Fixed the cell reference - was using KP2 instead of LP2
    ws4.merge_cells('LP2:MF2')
    ws4['LP2'] = "Organ Donor Medical Expenses"
    ws4['LP2'].font = black_font
    ws4['LP2'].alignment = center_alignment
    ws4['LP2'].number_format = '@'
    ws4.merge_cells('MG2:MJ2')
    ws4['MG2'] = "Personal Accident Cover"
    ws4['MG2'].font = black_font
    ws4['MG2'].alignment = center_alignment
    ws4['MG2'].number_format = '@'
    ws4.merge_cells('MK2:MM2')
    ws4['MK2'] = "Pre Existing Disease Benefit"
    ws4['MK2'].font = black_font
    ws4['MK2'].alignment = center_alignment
    ws4['MK2'].number_format = '@'
    ws4.merge_cells('MN2:NJ2')
    ws4['MN2'] = "Psychiatric Cover"
    ws4['MN2'].font = black_font
    ws4['MN2'].alignment = center_alignment
    ws4['MN2'].number_format = '@'
    ws4.merge_cells('NK2:OC2')
    ws4['NK2'] = "Recovery Benefit"
    ws4['NK2'].font = black_font
    ws4['NK2'].alignment = center_alignment
    ws4['NK2'].number_format = '@'
    ws4.merge_cells('OD2:OS2')
    ws4['OD2'] = "Referral Hospital Care"
    ws4['OD2'].font = black_font
    ws4['OD2'].alignment = center_alignment
    ws4['OD2'].number_format = '@'
    ws4.merge_cells('OT2:PK2')
    ws4['OT2'] = "Surgical Benefit"
    ws4['OT2'].font = black_font
    ws4['OT2'].alignment = center_alignment
    ws4['OT2'].number_format = '@'
    ws4.merge_cells('PL2:PO2')
    ws4['PL2'] = "Top Up Cover"
    ws4['PL2'].font = black_font
    ws4['PL2'].alignment = center_alignment
    ws4['PL2'].number_format = '@'
    ws4.merge_cells('PP2:QF2')
    ws4['PP2'] = "Vaccination/Immunization Cover"
    ws4['PP2'].font = black_font
    ws4['PP2'].alignment = center_alignment
    ws4['PP2'].number_format = '@'  

    column_headers_row3 = ["Number of Trips", "Sum Insured", "% Limit Applicable On", "Limit Percentage", "Limit Amount", "Applicability",
    "Valid from last consultation", "Consultation days", "Valid from date of discharge", "Discharge days",
    "Relationship", "Age", "Metro", "Provider Type", "Hospital Type", "Relationship", "Age From", "Age To", "City",
    "Provider Type", "Hospital Type", "Limit Applicable On", "% Applicable", "Amount Applicable", "Applicable Limit",
    "Over And Above Policy Sum Insured?", "Relationship", "Age", "Metro", "Provider Type", "Hospital Type",
    "Relationship", "Age From", "Age To", "City", "Provider Type", "Hospital Type", "Limit Applicable On",
    "% Applicable", "Amount Applicable", "Applicable Limit", "Max Days Per Illness", "Min Days Per Illness",
    "Over And Above Policy Sum Insured?", "Minimum LOS in days", "Applicable From", "Sum Insured", "Benefit Amount",
    "Over And Above Policy Sum Insured?", "Survival Period Applicable?", "Survival Period Applicable From",
    "Number of Days", "Relationship", "Age", "Metro", "Provider Type", "Hospital Type", "Relationship", "Age From",
    "Age To", "City", "Provider Type", "Hospital Type", "Limit Applicable On", "% Applicable", "Amount Applicable",
    "Applicable Limit", "Minimum LOS in days", "Over And Above Policy Sum Insured?", "Max Days Per Policy year",
    "Max Days Per Illness", "Fixed limt", "Sum Insured", "Threshold", "% Limit Applicable On", "Limit Percentage",
    "Limit Amount", "Applicability", "Open range", "Threshold SI", "Daily Limit Range From", "Daily Limit Range To",
    "Over And Above Policy Sum Insured?", "Sum Insured", "% Limit Applicable On", "Limit Percentage", "Limit Amount",
    "Applicability", "Relationship", "Age", "Metro", "Provider Type", "Hospital Type", "Relationship", "Age From",
    "Age To", "City", "Provider Type", "Hospital Type", "Limit Applicable On", "% Applicable", "Amount Applicable",
    "Applicable Limit", "Over And Above Policy Sum Insured?", "Applicable On?", "Is Doctor & Nursing Charges Combined?",
    "% Limit Applicable On", "Limit Percentage", "Limit Amount", "Applicability", "No of days Allowed",
    "Over And Above Sum Insured?", "Sum Insured", "% Limit Applicable On", "Limit Percentage", "Limit Amount",
    "Limit per child", "Applicability", "Over And Above Sum Insured?", "Relationship", "Age", "Metro", "Provider Type",
    "Hospital Type", "Relationship", "Age From", "Age To", "City", "Provider Type", "Hospital Type",
    "Limit Applicable On", "% Applicable", "Amount Applicable", "Applicable Limit", "Over And Above Sum Insured?",
    "Relationship", "Age", "Metro", "Provider Type", "Hospital Type", "Relationship", "Age From", "Age To", "City",
    "Provider Type", "Hospital Type", "Limit Applicable On", "% Applicable", "Amount Applicable", "Applicable Limit",
    "Over And Above Policy Sum Insured?", "Select Type", "% Limit Applicable On", "Limit Percentage", "Limit Amount",
    "Applicability", "Relationship", "Age", "Metro", "Provider Type", "Hospital Type", "Relationship", "Age From",
    "Age To", "City", "Provider Type", "Hospital Type", "Limit Applicable On", "% Applicable", "Amount Applicable",
    "Applicable Limit", "Claims Free Option Applicable", "Over And Above Policy Sum Insured?",
    "Benefit Applicability after policy years", "Frequency Of Health Check-up", "Relationship", "Age", "Metro",
    "Provider Type", "Hospital Type", "Relationship", "Age From", "Age To", "City", "Provider Type", "Hospital Type",
    "Limit Applicable On", "% Applicable", "Amount Applicable", "Applicable Limit", "Over And Above Policy Sum Insured?",
    "Minimum LOS in days", "No. of Hospital Beds", "SI Amount", "Minimum No. Days", "Maximum No. Days", "Sum Insured",
    "Benefit Amount", "Action", "Over And Above Policy Sum Insured?", "Time Access(In Days)", "Relationship", "Age",
    "Metro", "Provider Type", "Hospital Type", "Relationship", "Age From", "Age To", "City", "Provider Type",
    "Hospital Type", "Limit Applicable On", "% Applicable", "Amount Applicable", "Applicable Limit",
    "Over And Above Policy Sum Insured?", "Sum Insured", "% Limit Applicable On", "Limit Percentage", "Limit Amount",
    "Applicability", "Action", "Family Level Limits", "No of opinion Allowed", "Sum Insured", "% Limit Applicable On",
    "Limit Percentage", "Limit Amount", "Applicability", "Action", "Member Level Limits", "No of opinion Allowed",
    "Sum Insured", "% Limit Applicable On", "Limit Percentage", "Limit Amount", "Applicability", "Action",
    "Illness Applicability?", "Over And Above Policy Sum Insured?", "Relationship", "Age", "Metro", "Provider Type",
    "Hospital Type", "Relationship", "Age From", "Age To", "City", "Provider Type", "Hospital Type",
    "Limit Applicable On", "% Applicable", "Amount Applicable", "Applicable Limit", "Action",
    "Over And Above Policy Sum Insured?", "Aggregate Limit", "Sum Insured", "% Limit Applicable On",
    "Limit Percentage", "Limit Amount", "OP Treatment Limit", "Applicability", "Action",
    "Is Inclusive of Dental & Optical", "Limit Applicable On", "Individual", "SI Amount", "Sum Insured",
    "Benefit Applicable For", "% Limit Applicable On", "Limit Percentage", "Limit Amount", "Applicability", "Action",
    "Combined", "Sum Insured", "% Limit Applicable On", "Limit Percentage", "Limit Amount", "Applicability", "Action",
    "Relationship", "Age", "Metro", "Provider Type", "Hospital Type", "Relationship", "Age From", "Age To", "City",
    "Provider Type", "Hospital Type", "Limit Applicable On", "% Applicable", "Amount Applicable", "Applicable Limit",
    "Action", "Is Inclusive of Implant(Glass/Lens)", "Over And Above Policy Sum Insured?", "Select Type", "Lens",
    "Sum Insured", "% Limit Applicable On", "Limit Percentage", "Limit Amount", "Implant Applicable", "Implant Amount",
    "Applicability", "Action", "Glass", "Sum Insured", "% Limit Applicable On", "Limit Percentage", "Limit Amount",
    "Implant Applicable", "Implant Amount", "Applicability", "Action", "Sum Insured", "% Limit Applicable On",
    "Limit Percentage", "Limit Amount", "Applicability", "Action", "Over And Above Policy Sum Insured?",
    "Relationship", "Age", "Metro", "Provider Type", "Hospital Type", "Relationship", "Age From","Age To", "City",
    "Provider Type","Hospital Type","Limit Applicable On","% Applicable","Amount Applicable","Applicable Limit","Action","Over And Above Policy Sum Insured?","Sum Insured","PA Sum Insured","Action","Member Waiting Period","Family Waiting Period","Policy Waiting Period","Over And Above Policy Sum Insured?",
    "Sum Insured","% Limit Applicable On","Limit Percentage","Limit Amount","Applicability", "Action", "Relationship", "Age", "Metro", "Provider Type", "Hospital Type",
    "Relationship", "Age From", "Age To", "City", "Provider Type", "Hospital Type", "Limit Applicable On",
    "% Applicable", "Amount Applicable", "Applicable Limit", "Action", "Recovery Period",
    "Over And Above Policy Sum Insured?", "Applicable From", "Relationship", "Age", "Metro", "Provider Type",
    "Hospital Type", "Relationship", "Age From", "Age To", "City", "Provider Type", "Hospital Type",
    "Limit Applicable On", "% Applicable", "Amount Applicable", "Applicable Limit", "Action", "Relationship", "Age",
    "Metro", "Provider Type", "Hospital Type", "Relationship", "Age From", "Age To", "City", "Provider Type",
    "Hospital Type", "Limit Applicable On", "% Applicable", "Amount Applicable", "Applicable Limit", "Action",
    "Over And Above Policy Sum Insured?", "Surgeries Covered", "Relationship", "Age", "Metro", "Provider Type",
    "Hospital Type", "Relationship", "Age From", "Age To", "City", "Provider Type", "Hospital Type",
    "Limit Applicable On", "% Applicable", "Amount Applicable", "Applicable Limit", "Action",
    "Over And Above Policy Sum Insured?", "Sum Insured", "Top Up Sum Insured", "Action",
    "Over And Above Policy Sum Insured?", "Relationship", "Age", "Metro", "Provider Type", "Hospital Type",
    "Relationship", "Age From", "Age To", "City", "Provider Type", "Hospital Type", "Limit Applicable On",
    "% Applicable", "Amount Applicable", "Applicable Limit", "Action"]

    for col, header in enumerate(column_headers_row3, 1):
            cell = ws4.cell(row=3, column=col, value=header)
            cell.font = black_font
            cell.alignment = left_alignment

    if AddonCoverages_data:
        # FLEXIBLE COLUMN MAPPING: You can decide which data goes in which column
        # Just modify the column numbers and field names as needed
        row_idx = 4  # Default row for data
        
        # Get all available data from different coverage types
        all_data = {}
        
        # Collect Critical Illness data
        if "Critical Illness" in AddonCoverages_data and AddonCoverages_data["Critical Illness"]:
            ci_data = AddonCoverages_data["Critical Illness"][0] if isinstance(AddonCoverages_data["Critical Illness"], list) and len(AddonCoverages_data["Critical Illness"]) > 0 else AddonCoverages_data["Critical Illness"]
            all_data.update(ci_data)
        
        # Collect Ambulance Cover data
        if "Ambulance Cover" in AddonCoverages_data and AddonCoverages_data["Ambulance Cover"]:
            amb_data = AddonCoverages_data["Ambulance Cover"][0] if isinstance(AddonCoverages_data["Ambulance Cover"], list) and len(AddonCoverages_data["Ambulance Cover"]) > 0 else AddonCoverages_data["Ambulance Cover"]
            all_data.update(amb_data)
        
        # Collect Convalescence Benefit data
        if "Convalescence Benefit" in AddonCoverages_data and AddonCoverages_data["Convalescence Benefit"]:
            conv_data = AddonCoverages_data["Convalescence Benefit"][0] if isinstance(AddonCoverages_data["Convalescence Benefit"], list) and len(AddonCoverages_data["Convalescence Benefit"]) > 0 else AddonCoverages_data["Convalescence Benefit"]
            all_data.update(conv_data)
        
        # Collect Daily Cash Cover data
        if "Daily Cash Cover" in AddonCoverages_data and AddonCoverages_data["Daily Cash Cover"]:
            daily_cash_data = AddonCoverages_data["Daily Cash Cover"][0] if isinstance(AddonCoverages_data["Daily Cash Cover"], list) and len(AddonCoverages_data["Daily Cash Cover"]) > 0 else AddonCoverages_data["Daily Cash Cover"]
            all_data.update(daily_cash_data)
        
        # Collect Home Nursing Allowance data
        if "Home Nursing Allowance" in AddonCoverages_data and AddonCoverages_data["Home Nursing Allowance"]:
            home_nursing_data = AddonCoverages_data["Home Nursing Allowance"][0] if isinstance(AddonCoverages_data["Home Nursing Allowance"], list) and len(AddonCoverages_data["Home Nursing Allowance"]) > 0 else AddonCoverages_data["Home Nursing Allowance"]
            all_data.update(home_nursing_data)
        
        # NOW YOU CAN MAP ANY FIELD TO ANY COLUMN
        # Format: ws4.cell(row=row_idx, column=COLUMN_NUMBER, value=all_data.get("FIELD_NAME", ""))
        #Ambulance Cover
        try:
            ws4.cell(row=row_idx, column=1, value=float(all_data.get("Ambulance_Number_of_Trips", ""))).number_format = "#,##0"
        except (ValueError, TypeError):
            ws4.cell(row=row_idx, column=1, value=all_data.get("Ambulance_Number_of_Trips", "")).number_format = "#,##0"
        ws4.cell(row=row_idx, column=2, value=all_data.get("Ambulance_Sum_Insured", "")).number_format = "#,##0"
        ws4.cell(row=row_idx, column=3, value=all_data.get("Ambulance_Limit_Applicable_On", "")).number_format = "@"
        ws4.cell(row=row_idx, column=4, value=all_data.get("Ambulance_Limit_Percentage", "")).number_format = "0.00%"
        try:
            ws4.cell(row=row_idx, column=5, value=float(all_data.get("Ambulance_Limit_Amount", ""))).number_format = "#,##0"
        except (ValueError, TypeError):
            ws4.cell(row=row_idx, column=5, value=all_data.get("Ambulance_Limit_Amount", "")).number_format = "#,##0"
        ws4.cell(row=row_idx, column=6, value=all_data.get("Ambulance_Applicability", "")).number_format = "@"
        
        #Convalescence Benefit
        try:
            ws4.cell(row=row_idx, column=45, value=float(all_data.get("Convalescence_Minimum_LOS_in_days", ""))).number_format = "#,##0"
        except (ValueError, TypeError):
            ws4.cell(row=row_idx, column=45, value=all_data.get("Convalescence_Minimum_LOS_in_days", "")).number_format = "#,##0"
        ws4.cell(row=row_idx, column=46, value=all_data.get("Convalescence_Applicable_From", ""))
        ws4.cell(row=row_idx, column=47, value=all_data.get("Convalescence_Sum_Insured", ""))
        try:
            ws4.cell(row=row_idx, column=48, value=float(all_data.get("Convalescence_Benefit_Amount", ""))).number_format = "#,##0"
        except (ValueError, TypeError):
            ws4.cell(row=row_idx, column=48, value=all_data.get("Convalescence_Benefit_Amount", "")).number_format = "#,##0"
        
        
        
        # Critical Illness - Check column 49 for yes/no logic
        column_49_value = all_data.get("Over And Above Policy Sum Insured?", "")
        ws4.cell(row=row_idx, column=49, value=column_49_value).number_format = "@"
        
        # Check if column 49 contains "yes" or "no" (case insensitive)
        if column_49_value and str(column_49_value).lower() in ["yes", "no"]:
            # If yes/no found, populate columns 49, 50, 64, 65, 66, 67, 68 with actual values
            ws4.cell(row=row_idx, column=50, value=all_data.get("Survival Period Applicable1?", "Yes")).number_format = "@"
            try:
                ws4.cell(row=row_idx, column=64, value=float(all_data.get("Applicable_Limit_default", "500000"))).number_format = "#,##0"
            except (ValueError, TypeError):
                ws4.cell(row=row_idx, column=64, value=all_data.get("Applicable_Limit_default", "500000")).number_format = "#,##0"
            try:
                ws4.cell(row=row_idx, column=65, value=float(all_data.get("Maximum Limit Percentage", "0"))).number_format = "0.00%"
            except (ValueError, TypeError):
                ws4.cell(row=row_idx, column=65, value=all_data.get("Maximum Limit Percentage", "0")).number_format = "0.00%"
            ws4.cell(row=row_idx, column=66, value=all_data.get("Maximum Limit", "")).number_format = "#,##0"
            ws4.cell(row=row_idx, column=67, value=all_data.get("critical_Applicability", "Lower")).number_format = "@"
            
        else:
            # If no yes/no value, leave columns 49, 50, 64, 65, 66, 67, 68 blank
            ws4.cell(row=row_idx, column=50, value="")
            ws4.cell(row=row_idx, column=64, value="")
            ws4.cell(row=row_idx, column=65, value="")
            ws4.cell(row=row_idx, column=66, value="")
            ws4.cell(row=row_idx, column=67, value="")
            

        # Daily Cash Cover mappings - Check column 69 for yes/no logic
        column_69_value = all_data.get("DailyCash_Over_And_Above_Policy_Sum_Insured", "")
        ws4.cell(row=row_idx, column=69, value=column_69_value).number_format = "@"
        
        # Check if column 69 contains "yes" or "no" (case insensitive)
        if column_69_value and str(column_69_value).lower() in ["yes", "no"]:
            # If yes/no found, populate columns 68, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82 with actual values
            try:
                ws4.cell(row=row_idx, column=68, value=float(all_data.get("DailyCash_Minimum_LOS_in_days", ""))).number_format = "#,##0"
            except (ValueError, TypeError):
                ws4.cell(row=row_idx, column=68, value=all_data.get("DailyCash_Minimum_LOS_in_days", "")).number_format = "#,##0"
            try:
                ws4.cell(row=row_idx, column=70, value=float(all_data.get("DailyCash_Max_Days_Per_Policy_year", ""))).number_format = "#,##0"
            except (ValueError, TypeError):
                ws4.cell(row=row_idx, column=70, value=all_data.get("DailyCash_Max_Days_Per_Policy_year", "")).number_format = "#,##0"
            try:
                ws4.cell(row=row_idx, column=71, value=float(all_data.get("DailyCash_Max_Days_Per_Illness", ""))).number_format = "#,##0"
            except (ValueError, TypeError):
                ws4.cell(row=row_idx, column=71, value=all_data.get("DailyCash_Max_Days_Per_Illness", "")).number_format = "#,##0"
            ws4.cell(row=row_idx, column=72, value=all_data.get("DailyCash_Fixed_limit", "")).number_format = "@"
            ws4.cell(row=row_idx, column=73, value=all_data.get("DailyCash_Sum_Insured_default", "500000"))
            ws4.cell(row=row_idx, column=74, value=all_data.get("DailyCash_Threshold1", ""))
            ws4.cell(row=row_idx, column=75, value=all_data.get("DailyCash_Sum Insured", "Yes")).number_format = "@"
            try:
                ws4.cell(row=row_idx, column=76, value=float(all_data.get("DailyCash_Daily_cash_percentage", "")))
            except (ValueError, TypeError):
                ws4.cell(row=row_idx, column=76, value=all_data.get("DailyCash_Daily_cash_percentage", ""))
            try:
                ws4.cell(row=row_idx, column=77, value=float(all_data.get("DailyCash_Limit_Amount", ""))).number_format = "#,##0"
            except (ValueError, TypeError):
                ws4.cell(row=row_idx, column=77, value=all_data.get("DailyCash_Limit_Amount", "")).number_format = "#,##0"
            ws4.cell(row=row_idx, column=78, value=all_data.get("DailyCash_Applicability", "Lower")).number_format = "@"
            ws4.cell(row=row_idx, column=79, value=all_data.get("DailyCash_Open_range", "")).number_format = "@"
            ws4.cell(row=row_idx, column=80, value=all_data.get("DailyCash_Waiting_Period_Days 1", ""))
            try:
                ws4.cell(row=row_idx, column=81, value=float(all_data.get("DailyCash_Daily_Limit_Range_From", ""))).number_format = "#,##0"
            except (ValueError, TypeError):
                ws4.cell(row=row_idx, column=81, value=all_data.get("DailyCash_Daily_Limit_Range_From", "")).number_format = "#,##0"
            try:
                ws4.cell(row=row_idx, column=82, value=float(all_data.get("DailyCash_Daily_Limit_Range_To", ""))).number_format = "#,##0"
            except (ValueError, TypeError):
                ws4.cell(row=row_idx, column=82, value=all_data.get("DailyCash_Daily_Limit_Range_To", "")).number_format = "#,##0"
        else:
            # If no yes/no value, leave columns 68, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82 blank
            ws4.cell(row=row_idx, column=68, value="")
            ws4.cell(row=row_idx, column=70, value="")
            ws4.cell(row=row_idx, column=71, value="")
            ws4.cell(row=row_idx, column=72, value="")
            ws4.cell(row=row_idx, column=73, value="")
            ws4.cell(row=row_idx, column=74, value="")
            ws4.cell(row=row_idx, column=75, value="")
            ws4.cell(row=row_idx, column=76, value="")
            ws4.cell(row=row_idx, column=77, value="")
            ws4.cell(row=row_idx, column=78, value="")
            ws4.cell(row=row_idx, column=79, value="")
            ws4.cell(row=row_idx, column=80, value="")
            ws4.cell(row=row_idx, column=81, value="")
            ws4.cell(row=row_idx, column=82, value="")

        # Applicability of Doctor's Home Visit & Nursing Charges
        ws4.cell(row=row_idx, column=105, value=all_data.get("Nursing_Applicable_On", "")).number_format = "@"
        ws4.cell(row=row_idx, column=106, value=all_data.get("Nursing_Doctor_Nursing_Combined", "")).number_format = "@"
        ws4.cell(row=row_idx, column=107, value=all_data.get("Nursing_Limit_Applicable_On", "")).number_format = "@"
        ws4.cell(row=row_idx, column=108, value=all_data.get("Nursing_Limit_Percentage", ""))
        try:
            ws4.cell(row=row_idx, column=109, value=float(all_data.get("Nursing_Limit_Amount", ""))).number_format = "#,##0"
        except (ValueError, TypeError):
            ws4.cell(row=row_idx, column=109, value=all_data.get("Nursing_Limit_Amount", "")).number_format = "#,##0"
        ws4.cell(row=row_idx, column=110, value=all_data.get("Nursing_Applicability", ""))
        try:
            ws4.cell(row=row_idx, column=111, value=float(all_data.get("Nursing_Days_Allowed", ""))).number_format = "#,##0"
        except (ValueError, TypeError):
            ws4.cell(row=row_idx, column=111, value=all_data.get("Nursing_Days_Allowed", "")).number_format = "#,##0"
        # ADD MORE MAPPINGS HERE - Just copy and modify:
        # ws4.cell(row=row_idx, column=16, value=all_data.get("Your Field Name", ""))
        # ws4.cell(row=row_idx, column=17, value=all_data.get("Another Field", ""))
        # ws4.cell(row=row_idx, column=18, value=all_data.get("Third Field", ""))
        # ... and so on for any column you want
        
        # Available fields you can use:
        # Critical Illness: "Over And Above Policy Sum Insured?", "Survival Period Applicable?", "Applicable Limit", "Sum Insured Per Person", "Maximum Limit", "Survival Period", "Maximum Limit Percentage"
        # Ambulance Cover: "Sum Insured", "Number of Trips", "% Limit Applicable On", "Limit Amount", "Applicability", "Limit Percentage"
        # Convalescence Benefit: "Sum Insured", "Minimum LOS in days", "Applicable From", "Benefit Amount"
        # Daily Cash Cover: "DailyCash_Over_And_Above_Policy_Sum_Insured", "DailyCash_Max_Days_Per_Policy_year", "DailyCash_Max_Days_Per_Illness", "DailyCash_Fixed_limit", "DailyCash_Sum_Insured", "DailyCash_Threshold", "DailyCash_Limit_Amount", "DailyCash_Daily_Cash_Amount", "DailyCash_Daily_cash_percentage", "DailyCash_Minimum_Hospitalization_Days", "DailyCash_Minimum_LOS_in_days", "DailyCash_Maximum_Days_Per_Person", "DailyCash_Waiting_Period_Days", "DailyCash_Maternity_Exclusion", "DailyCash_First_Days_Exclusion", "DailyCash_Open_range"

    
    # Add data rows for sheet 4 (starting from row 4) with borders
    for row in range(4, 5):  # Row 4 (you can extend this range as needed)
        for col in range(1, 449):  # A to QF (1-448, so range needs to be 1-449)
            cell = ws4.cell(row=row, column=col)
            cell.border = thin_border

    # Set row heights for sheet 4
    ws4.row_dimensions[1].height = 18
    ws4.row_dimensions[2].height = 18
    ws4.row_dimensions[3].height = 35

    # Set column widths for sheet 4 - Further increased for better visibility
    column_widths_sheet4 = [35] * 448  # Create a list of 448 widths all set to 35 (A to QF) - further increased from 25
    
    for col, width in enumerate(column_widths_sheet4, 1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    return wb
